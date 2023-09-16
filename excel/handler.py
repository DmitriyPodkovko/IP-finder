import os
from typing import List, Tuple, Any

import openpyxl
import logging

from openpyxl.styles import NamedStyle

from config.settings import EXCEL_ROW_COLUMN
from config.settings import DESTINATION_NUMBER
from config.settings import EXCEL_OUTPUT_FILE_PREFIX
from config.settings import RESULT_DIRECTORY


class ExcelHandler:
    def __init__(self, xlsx_file) -> None:
        self._xlsx_file = xlsx_file
        self._cell_style = None
        self._new_column_idx = None
        self._xlsx_output_file = None
        self._current_row = EXCEL_ROW_COLUMN.get('default').get('Start_row')

    def get_ip_list_from_xlsx_file(self) -> list[tuple[Any, Any, Any, Any, Any]] | list[Any]:
        try:
            ip_list = []
            workbook = openpyxl.load_workbook(self._xlsx_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=EXCEL_ROW_COLUMN.get('default').get('Start_row'), values_only=True):
                ip_tuple = (
                    # row[EXCEL_ROW_COLUMN.get('default').get('SKP')],
                    row[EXCEL_ROW_COLUMN.get('default').get('IP_DST')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Port_DST')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Date')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Time')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Provider')]
                )
                ip_list.append(ip_tuple)
                logging.info(f'IP_DST, Port_DST, Date, Time, Provider: {ip_tuple}')
            workbook.close()
            return ip_list
        except Exception as e:
            logging.error(f'Error input .xlsx file:\n {self._xlsx_file} ({str(e)})')
            return []

    def create_output_xlsx_file(self) -> None:
        try:
            workbook = openpyxl.load_workbook(self._xlsx_file)
            sheet = workbook.active
            self._cell_style = NamedStyle(name='cell_style')
            self._cell_style.font = sheet.cell(row=1, column=1).font.copy()
            self._cell_style.border = sheet.cell(row=1, column=1).border.copy()
            # custom_style = NamedStyle(name="custom_style")
            # custom_style.font = sheet.cell(row=1, column=1).font.copy()
            # custom_style.border = sheet.cell(row=1, column=1).border.copy()
            # logging.info(f'{self._cell_style}')
            self._new_column_idx = sheet.max_column + 1
            sheet.cell(row=1, column=self._new_column_idx, value=DESTINATION_NUMBER)
            file_name, file_extension = os.path.splitext(self._xlsx_file.name)
            # logging.info(f'{os.getcwd()}')
            # logging.info(f'{file_name}  {file_extension}')
            # logging.info(f'{RESULT_DIRECTORY}')
            if not os.path.exists(RESULT_DIRECTORY):
                os.makedirs(RESULT_DIRECTORY)
            self._xlsx_output_file = f'{RESULT_DIRECTORY}/{file_name}{EXCEL_OUTPUT_FILE_PREFIX}.xlsx'
            workbook.save(self._xlsx_output_file)
            # workbook.save(self._xlsx_file.name.replace('.xlsx', EXCEL_OUTPUT_FILE_PREFIX + '.xlsx'))
            workbook.close()
            logging.info(f'CREATED: {self._xlsx_output_file}')
        except Exception as e:
            logging.error(f'Error creating .xlsx file:\n {self._xlsx_file} ({str(e)})')

    def save_result_to_output_xlsx_file(self, dst_set: set) -> None:
        try:
            workbook = openpyxl.load_workbook(self._xlsx_output_file)
            sheet = workbook.active
            next_column_idx = self._new_column_idx
            for i, value in enumerate(dst_set, start=self._current_row):
                # logging.info(f'next_column_idx={next_column_idx} value={value}')
                new_cell = sheet.cell(row=self._current_row, column=next_column_idx, value=value)
                new_cell.style = self._cell_style
                next_column_idx += 1
                # sheet.cell(row=1, column=next_column_idx, value=DESTINATION_NUMBER)
                workbook.save(self._xlsx_output_file)
            workbook.save(self._xlsx_output_file)
            workbook.close()
            self._current_row += 1
            logging.info(f'SAVED: {self._xlsx_output_file}')
        except Exception as e:
            logging.error(f'Error saving .xlsx:\n {self._xlsx_file} ({str(e)})')

    # def save_output_xlsx_file(self, dst_set: set) -> None:
    #     try:
    #         logging.info(f'{self._xlsx_file}')
    #         logging.info(f'{dst_set}')
    #         workbook = openpyxl.load_workbook(self._xlsx_file)
    #         sheet = workbook.active
    #         new_column_idx = sheet.max_column + 1
    #         sheet.cell(row=1, column=new_column_idx, value=DESTINATION_NUMBER)
    #         for row_idx, value in enumerate(dst_set, start=EXCEL_ROW_COLUMN.get('default').get('Start_row')):
    #             sheet.cell(row=row_idx, column=new_column_idx, value=value)
    #         # logging.info(f'{dst_set}')
    #     except Exception as e:
    #         logging.error(f'Error output .xlsx file:\n {self._xlsx_file} ({str(e)})')
